"""Tests for output writers (CSV, JSON, JSONL, XLSX, Statistics)."""

from __future__ import annotations

import json

import pytest

from matches import PiiMatch
from core.exceptions import OutputError
from core.writers import (
    CsvWriter,
    JsonWriter,
    JsonlWriter,
    PrivacyStatisticsWriter,
    create_output_writer,
)


def test_csv_writer_write_and_finalize(tmp_path):
    """Test CsvWriter writes matches and finalizes correctly."""
    out = tmp_path / "findings.csv"
    writer = CsvWriter(str(out))

    writer.write_match(
        PiiMatch(
            text="test@example.com",
            file="/tmp/a.txt",
            type="REGEX_EMAIL",
            ner_score=None,
            engine="regex",
            metadata={},
        )
    )
    writer.finalize()

    content = out.read_text(encoding="utf-8")
    assert "Match" in content
    assert "test@example.com" in content
    assert writer.supports_streaming is True


def test_json_writer_write_and_finalize(tmp_path):
    """Test JsonWriter writes matches and finalizes correctly."""
    out = tmp_path / "findings.json"
    writer = JsonWriter(str(out))

    writer.write_match(
        PiiMatch(
            text="John Doe",
            file="/tmp/a.txt",
            type="NER_PERSON",
            ner_score=0.9,
            engine="gliner",
            metadata={},
        )
    )
    writer.finalize(metadata={"scan_id": "123"})

    data = json.loads(out.read_text(encoding="utf-8"))
    assert "findings" in data
    assert len(data["findings"]) == 1
    assert data["findings"][0]["text"] == "John Doe"
    assert data["metadata"]["scan_id"] == "123"
    assert writer.supports_streaming is False


def test_jsonl_writer_streams_and_appends_metadata(tmp_path):
    out = tmp_path / "findings.jsonl"
    writer = create_output_writer("jsonl", str(out))

    writer.write_match(
        PiiMatch(
            text="user@example.com",
            file="/tmp/a.txt",
            type="REGEX_EMAIL",
            ner_score=None,
            engine="regex",
            metadata={},
        )
    )
    writer.finalize(metadata={"k": "v"})

    lines = out.read_text(encoding="utf-8").splitlines()
    assert len(lines) == 2
    first = json.loads(lines[0])
    assert first["text"] == "user@example.com"
    assert first["type"] == "REGEX_EMAIL"
    last = json.loads(lines[1])
    assert last["_metadata"]["k"] == "v"


def test_xlsx_writer_streams_rows(tmp_path):
    out = tmp_path / "findings.xlsx"
    writer = create_output_writer("xlsx", str(out))

    writer.write_match(
        PiiMatch(
            text="John Doe",
            file="/tmp/a.txt",
            type="NER_PERSON",
            ner_score=0.9,
            engine="gliner",
            metadata={},
        )
    )
    writer.finalize(metadata={"scan": "id"})

    import openpyxl

    wb = openpyxl.load_workbook(str(out), read_only=True, data_only=True)
    assert "Findings" in wb.sheetnames
    ws = wb["Findings"]
    rows = list(ws.iter_rows(values_only=True))
    # header + one row
    assert rows[0] == ("Match", "File", "Type", "Score", "Engine")
    assert rows[1][0] == "John Doe"
    assert "Metadata" in wb.sheetnames


def test_create_output_writer_csv(tmp_path):
    """Test create_output_writer returns CsvWriter for csv format."""
    out = tmp_path / "out.csv"
    writer = create_output_writer("csv", str(out))
    assert isinstance(writer, CsvWriter)


def test_create_output_writer_json(tmp_path):
    """Test create_output_writer returns JsonWriter for json format."""
    out = tmp_path / "out.json"
    writer = create_output_writer("json", str(out))
    assert isinstance(writer, JsonWriter)


def test_create_output_writer_jsonl(tmp_path):
    """Test create_output_writer returns JsonlWriter for jsonl format."""
    out = tmp_path / "out.jsonl"
    writer = create_output_writer("jsonl", str(out))
    assert isinstance(writer, JsonlWriter)


def test_create_output_writer_defaults_to_csv(tmp_path):
    """Test create_output_writer defaults to CSV for unknown format."""
    out = tmp_path / "out.csv"
    writer = create_output_writer("unknown", str(out))
    assert isinstance(writer, CsvWriter)


def test_csv_writer_io_error(tmp_path):
    """Test CsvWriter raises OutputError on invalid path."""
    invalid_path = tmp_path / "nonexistent" / "nested" / "file.csv"
    with pytest.raises(OutputError) as exc_info:
        CsvWriter(str(invalid_path))
    assert "Failed to open output file" in str(exc_info.value)


def test_create_output_writer_statistics(tmp_path):
    """Test create_output_writer returns PrivacyStatisticsWriter for statistics format."""
    out = tmp_path / "stats.json"
    writer = create_output_writer("statistics", str(out))
    assert isinstance(writer, PrivacyStatisticsWriter)


def test_privacy_statistics_writer(tmp_path):
    """Test PrivacyStatisticsWriter writes aggregated statistics."""
    out = tmp_path / "stats.json"
    writer = PrivacyStatisticsWriter(str(out))
    writer.write_match(
        PiiMatch(
            text="test",
            file="/a.txt",
            type="REGEX",
            ner_score=None,
            engine="regex",
            metadata={},
        )
    )
    writer.finalize(
        metadata={
            "statistics": {
                "statistics_by_dimension": {"pii": {"count": 1}},
                "summary": {"total": 1},
            },
            "scan_metadata": {"path": "/test"},
        }
    )
    data = json.loads(out.read_text(encoding="utf-8"))
    assert "statistics_by_dimension" in data
    assert "summary" in data
    assert "metadata" in data


"""Output writers for PII findings."""

import abc
import csv
import html
import json
from typing import Optional, Any, TextIO

from matches import PiiMatch
from core.exceptions import OutputError


class OutputWriter(abc.ABC):
    """Abstract base class for output writers."""

    def __init__(self, file_path: str, include_header: bool = True):
        self.file_path = file_path
        self.include_header = include_header

    @abc.abstractmethod
    def write_match(self, match: PiiMatch) -> None:
        """Write a single match to the output."""
        pass

    @abc.abstractmethod
    def finalize(self, metadata: Optional[dict] = None) -> None:
        """Finalize the output (e.g. close file, write footer/metadata)."""
        pass

    @property
    @abc.abstractmethod
    def supports_streaming(self) -> bool:
        """Return True if the writer supports streaming (writing matches as they are found)."""
        pass

    # Optional method for backward compatibility
    def get_writer(self) -> Any:
        return None

    @property
    def file_handle(self) -> Optional[Any]:
        return None


class CsvWriter(OutputWriter):
    """Writes findings to a CSV file."""

    def __init__(self, file_path: str, include_header: bool = True):
        super().__init__(file_path, include_header)
        try:
            self._file = open(file_path, "w", newline="", encoding="utf-8")
            self._writer = csv.writer(self._file)
            if self.include_header:
                self._writer.writerow(
                    ["Match", "File", "Type", "Score", "Engine", "Severity"]
                )
        except IOError as e:
            raise OutputError(f"Failed to open output file: {e}")

    def write_match(self, match: PiiMatch) -> None:
        row = [
            match.text,
            match.file,
            match.type,
            match.ner_score,
            match.engine,
            match.severity,
        ]
        self._writer.writerow(row)

    def finalize(self, metadata: Optional[dict] = None) -> None:
        if self._file:
            self._file.close()
            self._file = None

    @property
    def supports_streaming(self) -> bool:
        return True

    def get_writer(self) -> csv.writer:
        return self._writer

    @property
    def file_handle(self) -> TextIO:
        return self._file


class JsonWriter(OutputWriter):
    """Writes findings to a JSON file."""

    def __init__(self, file_path: str, include_header: bool = True):
        super().__init__(file_path, include_header)
        self.matches: list[dict] = []

    def write_match(self, match: PiiMatch) -> None:
        # Convert match to dict
        match_dict = {
            "text": match.text,
            "file": match.file,
            "type": match.type,
            "score": match.ner_score,
            "engine": match.engine,
            "severity": match.severity,
            "metadata": match.metadata,
        }
        # Include context fields when available
        if match.context_before is not None:
            match_dict["context_before"] = match.context_before
        if match.context_after is not None:
            match_dict["context_after"] = match.context_after
        if match.char_offset is not None:
            match_dict["char_offset"] = match.char_offset
        self.matches.append(match_dict)

    def finalize(self, metadata: Optional[dict] = None) -> None:
        output_data = {"metadata": metadata or {}, "findings": self.matches}
        try:
            with open(self.file_path, "w", encoding="utf-8") as f:
                json.dump(output_data, f, indent=2, ensure_ascii=False)
        except IOError as e:
            raise OutputError(f"Failed to write JSON output: {e}")

    @property
    def supports_streaming(self) -> bool:
        return False


class JsonlWriter(OutputWriter):
    """Writes findings to a JSON Lines (JSONL) file.

    Each match is written as one JSON object per line for streaming and easy
    incremental processing. Metadata is appended as a final line with the key
    "_metadata".
    """

    def __init__(self, file_path: str, include_header: bool = True):
        super().__init__(file_path, include_header)
        try:
            self._file = open(file_path, "w", encoding="utf-8")
        except IOError as e:
            raise OutputError(f"Failed to open output file: {e}")

    def write_match(self, match: PiiMatch) -> None:
        payload = {
            "text": match.text,
            "file": match.file,
            "type": match.type,
            "score": match.ner_score,
            "engine": match.engine,
            "severity": match.severity,
            "metadata": match.metadata,
        }
        if match.context_before is not None:
            payload["context_before"] = match.context_before
        if match.context_after is not None:
            payload["context_after"] = match.context_after
        if match.char_offset is not None:
            payload["char_offset"] = match.char_offset
        self._file.write(json.dumps(payload, ensure_ascii=False) + "\n")

    def finalize(self, metadata: Optional[dict] = None) -> None:
        if metadata:
            self._file.write(
                json.dumps({"_metadata": metadata}, ensure_ascii=False) + "\n"
            )
        if self._file:
            self._file.close()
            self._file = None

    @property
    def supports_streaming(self) -> bool:
        return True

    @property
    def file_handle(self) -> TextIO:
        return self._file


class XlsxWriter(OutputWriter):
    """Writes findings to an Excel file (streaming, write-only workbook)."""

    def __init__(self, file_path: str, include_header: bool = True):
        super().__init__(file_path, include_header)
        try:
            import openpyxl
        except ImportError:
            raise OutputError(
                "openpyxl is required for XLSX output but is not installed."
            )

        # Use write_only mode to avoid holding all rows in memory.
        self._openpyxl = openpyxl
        self._wb = openpyxl.Workbook(write_only=True)
        self._ws = self._wb.create_sheet("Findings")
        if self.include_header:
            self._ws.append(["Match", "File", "Type", "Score", "Engine", "Severity"])

    def write_match(self, match: PiiMatch) -> None:
        self._ws.append(
            [
                match.text,
                match.file,
                match.type,
                match.ner_score,
                match.engine,
                match.severity,
            ]
        )

    def finalize(self, metadata: Optional[dict] = None) -> None:
        # Add metadata sheet
        if metadata:
            ws_meta = self._wb.create_sheet("Metadata")
            ws_meta.append(["Key", "Value"])
            # Flatten metadata if needed or just dump top level
            for k, v in metadata.items():
                if isinstance(v, (dict, list)):
                    v = json.dumps(v)
                ws_meta.append([k, v])

        try:
            self._wb.save(self.file_path)
        except IOError as e:
            raise OutputError(f"Failed to save Excel file: {e}")

    @property
    def supports_streaming(self) -> bool:
        return True


class PrivacyStatisticsWriter(OutputWriter):
    """Writes privacy-focused statistics to a JSON file.

    This writer generates aggregated statistics by privacy dimensions and
    detection modules without storing individual PII instances or file paths.
    """

    def __init__(self, file_path: str, include_header: bool = True):
        super().__init__(file_path, include_header)
        # This writer doesn't collect matches, but we need to implement write_match
        # for compatibility with the abstract base class
        self._match_count = 0

    def write_match(self, match: PiiMatch) -> None:
        """Write a match (for compatibility).

        Note: This writer doesn't actually store matches, but tracks count
        for validation purposes.

        Args:
            match: PiiMatch object (not stored, only counted)
        """
        self._match_count += 1

    def finalize(self, metadata: Optional[dict] = None) -> None:
        """Write aggregated statistics to JSON file.

        Args:
            metadata: Dictionary containing:
                - statistics: Aggregated statistics from StatisticsAggregator
                - scan_metadata: Scan metadata (start_time, end_time, etc.)
        """
        if metadata is None:
            metadata = {}

        # Extract statistics and scan metadata
        statistics = metadata.get("statistics", {})
        scan_metadata = metadata.get("scan_metadata", {})

        # Build output structure
        output_data = {
            "metadata": scan_metadata,
            "statistics_by_dimension": statistics.get("statistics_by_dimension", {}),
            "statistics_by_module": statistics.get("statistics_by_module", {}),
            "statistics_by_file_type": statistics.get("statistics_by_file_type", {}),
            "summary": statistics.get("summary", {}),
            "performance_metrics": metadata.get("performance_metrics", {}),
        }

        try:
            with open(self.file_path, "w", encoding="utf-8") as f:
                json.dump(output_data, f, indent=2, ensure_ascii=False)
        except IOError as e:
            raise OutputError(f"Failed to write statistics JSON output: {e}")

    @property
    def supports_streaming(self) -> bool:
        return False


class HtmlWriter(OutputWriter):
    """Writes findings to a self-contained HTML report with interactive dashboard."""

    def __init__(self, file_path: str, include_header: bool = True):
        super().__init__(file_path, include_header)
        self.matches: list[dict] = []

    def write_match(self, match: PiiMatch) -> None:
        self.matches.append({
            "text": match.text,
            "file": match.file,
            "type": match.type,
            "score": match.ner_score,
            "engine": match.engine,
            "severity": match.severity,
        })

    def finalize(self, metadata: Optional[dict] = None) -> None:
        meta = metadata or {}
        start_time = html.escape(str(meta.get("start_time", "N/A")))
        duration = html.escape(str(meta.get("duration", "N/A")))
        files_scanned = html.escape(str(meta.get("files_scanned", "N/A")))
        total_findings = len(self.matches)

        sev_counts = {"CRITICAL": 0, "HIGH": 0, "MEDIUM": 0, "LOW": 0}
        for m in self.matches:
            sev = m.get("severity", "LOW")
            if sev in sev_counts:
                sev_counts[sev] += 1

        # Build table rows
        rows_html = ""
        for m in self.matches:
            sev = html.escape(str(m.get("severity", "")))
            rows_html += (
                "<tr>"
                f"<td>{html.escape(str(m.get('file', '')))}</td>"
                f"<td>{html.escape(str(m.get('type', '')))}</td>"
                f"<td class=\"match-text\">{html.escape(str(m.get('text', '')))}</td>"
                f"<td>{html.escape(str(m.get('engine', '')))}</td>"
                f"<td><span class=\"badge sev-{sev}\" onclick=\"filterSev('{sev}')\">{sev}</span></td>"
                f"<td>{html.escape(str(m.get('score', '')))}</td>"
                "</tr>\n"
            )

        page = f"""<!DOCTYPE html>
<html lang="en"><head><meta charset="utf-8">
<title>PII Scan Report</title>
<style>
*{{box-sizing:border-box;margin:0;padding:0}}
body{{font-family:system-ui,sans-serif;background:#f5f6fa;color:#23272f;padding:20px}}
.header{{background:#23272f;color:#fff;padding:24px 32px;border-radius:8px;margin-bottom:20px}}
.header h1{{font-size:1.5rem;margin-bottom:8px}}
.meta span{{margin-right:18px;font-size:.9rem;opacity:.85}}
.cards{{display:flex;gap:12px;margin-bottom:20px;flex-wrap:wrap}}
.card{{flex:1;min-width:140px;padding:16px;border-radius:8px;text-align:center;color:#fff;font-weight:bold}}
.card .count{{font-size:2rem}}
.card .label{{font-size:.8rem;opacity:.9}}
.bg-crit{{background:#d32f2f}} .bg-high{{background:#ef6c00}}
.bg-med{{background:#f9a825}} .bg-low{{background:#388e3c}}
.controls{{margin-bottom:12px;display:flex;gap:10px;align-items:center;flex-wrap:wrap}}
#search{{padding:8px 12px;border:1px solid #ccc;border-radius:4px;width:280px;font-size:.9rem}}
#sevFilter{{padding:8px;border:1px solid #ccc;border-radius:4px;font-size:.9rem}}
table{{width:100%;border-collapse:collapse;background:#fff;border-radius:8px;overflow:hidden;box-shadow:0 1px 4px rgba(0,0,0,.08)}}
th{{background:#23272f;color:#fff;padding:10px 12px;cursor:pointer;user-select:none;font-size:.85rem;text-align:left}}
th:hover{{background:#3a3f4b}}
td{{padding:8px 12px;border-bottom:1px solid #eee;font-size:.85rem;max-width:300px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}}
tr:hover{{background:#f0f4ff}}
.badge{{padding:3px 10px;border-radius:12px;font-size:.75rem;cursor:pointer;color:#fff;display:inline-block}}
.sev-CRITICAL{{background:#d32f2f}} .sev-HIGH{{background:#ef6c00}}
.sev-MEDIUM{{background:#f9a825;color:#333}} .sev-LOW{{background:#388e3c}}
.match-text{{font-family:monospace;font-size:.8rem}}
</style></head><body>
<div class="header">
<h1>PII Detection Report</h1>
<div class="meta">
<span>Start: {start_time}</span>
<span>Duration: {duration}</span>
<span>Files scanned: {files_scanned}</span>
<span>Total findings: {total_findings}</span>
</div></div>
<div class="cards">
<div class="card bg-crit"><div class="count">{sev_counts['CRITICAL']}</div><div class="label">CRITICAL</div></div>
<div class="card bg-high"><div class="count">{sev_counts['HIGH']}</div><div class="label">HIGH</div></div>
<div class="card bg-med"><div class="count">{sev_counts['MEDIUM']}</div><div class="label">MEDIUM</div></div>
<div class="card bg-low"><div class="count">{sev_counts['LOW']}</div><div class="label">LOW</div></div>
</div>
<div class="controls">
<input id="search" placeholder="Search findings..." oninput="applyFilters()">
<select id="sevFilter" onchange="applyFilters()">
<option value="">All Severities</option>
<option value="CRITICAL">CRITICAL</option>
<option value="HIGH">HIGH</option>
<option value="MEDIUM">MEDIUM</option>
<option value="LOW">LOW</option>
</select>
</div>
<table><thead><tr>
<th onclick="sortTable(0)">File</th>
<th onclick="sortTable(1)">Type</th>
<th onclick="sortTable(2)">Match Text</th>
<th onclick="sortTable(3)">Engine</th>
<th onclick="sortTable(4)">Severity</th>
<th onclick="sortTable(5)">Score</th>
</tr></thead><tbody id="tbody">
{rows_html}</tbody></table>
<script>
let sortCol=-1,sortAsc=true;
function sortTable(c){{if(sortCol===c)sortAsc=!sortAsc;else{{sortCol=c;sortAsc=true}}
const tb=document.getElementById('tbody');
const rows=Array.from(tb.rows);
rows.sort((a,b)=>{{let va=a.cells[c].textContent,vb=b.cells[c].textContent;
if(c===5){{va=parseFloat(va)||0;vb=parseFloat(vb)||0;return sortAsc?va-vb:vb-va}}
return sortAsc?va.localeCompare(vb):vb.localeCompare(va)}});
rows.forEach(r=>tb.appendChild(r))}}
function applyFilters(){{const q=document.getElementById('search').value.toLowerCase();
const sev=document.getElementById('sevFilter').value;
const rows=document.getElementById('tbody').rows;
for(let r of rows){{const txt=r.textContent.toLowerCase();
const rSev=r.cells[4].textContent.trim();
r.style.display=(txt.includes(q)&&(!sev||rSev===sev))?'':'none'}}}}
function filterSev(s){{document.getElementById('sevFilter').value=s;applyFilters()}}
</script></body></html>"""

        try:
            with open(self.file_path, "w", encoding="utf-8") as f:
                f.write(page)
        except IOError as e:
            raise OutputError(f"Failed to write HTML output: {e}")

    @property
    def supports_streaming(self) -> bool:
        return False


class SarifWriter(OutputWriter):
    """Writes findings as a SARIF 2.1.0 JSON file for CI/CD integration."""

    _SEVERITY_MAP = {
        "CRITICAL": "error",
        "HIGH": "error",
        "MEDIUM": "warning",
        "LOW": "note",
    }

    def __init__(self, file_path: str, include_header: bool = True):
        super().__init__(file_path, include_header)
        self._matches: list[PiiMatch] = []

    def write_match(self, match: PiiMatch) -> None:
        self._matches.append(match)

    def finalize(self, metadata: Optional[dict] = None) -> None:
        # Build unique rules from PII types seen
        rule_ids: list[str] = []
        rule_index: dict[str, int] = {}
        for m in self._matches:
            if m.type not in rule_index:
                rule_index[m.type] = len(rule_ids)
                rule_ids.append(m.type)

        rules = [
            {"id": rid, "shortDescription": {"text": f"PII detected: {rid}"}}
            for rid in rule_ids
        ]

        # Build results
        results = []
        for m in self._matches:
            truncated = m.text[:100] if m.text else ""
            results.append({
                "ruleId": m.type,
                "ruleIndex": rule_index[m.type],
                "level": self._SEVERITY_MAP.get(m.severity, "note"),
                "message": {"text": truncated},
                "locations": [
                    {
                        "physicalLocation": {
                            "artifactLocation": {"uri": m.file},
                        }
                    }
                ],
            })

        sarif = {
            "$schema": "https://raw.githubusercontent.com/oasis-tcs/sarif-spec/main/sarif-2.1/schema/sarif-schema-2.1.0.json",
            "version": "2.1.0",
            "runs": [
                {
                    "tool": {
                        "driver": {
                            "name": "pii-toolkit",
                            "rules": rules,
                        }
                    },
                    "results": results,
                }
            ],
        }

        try:
            with open(self.file_path, "w", encoding="utf-8") as f:
                json.dump(sarif, f, indent=2, ensure_ascii=False)
        except IOError as e:
            raise OutputError(f"Failed to write SARIF output: {e}")

    @property
    def supports_streaming(self) -> bool:
        return False


def create_output_writer(
    output_format: str, file_path: str, include_header: bool = True
) -> OutputWriter:
    """Factory function to create the appropriate output writer."""
    if output_format == "json":
        return JsonWriter(file_path, include_header)
    elif output_format == "jsonl":
        return JsonlWriter(file_path, include_header)
    elif output_format == "xlsx":
        return XlsxWriter(file_path, include_header)
    elif output_format == "html":
        return HtmlWriter(file_path, include_header)
    elif output_format == "sarif":
        return SarifWriter(file_path, include_header)
    elif output_format == "statistics":
        return PrivacyStatisticsWriter(file_path, include_header)

    # Default to CSV
    return CsvWriter(file_path, include_header)

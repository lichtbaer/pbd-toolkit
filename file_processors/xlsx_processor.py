"""XLSX/XLS file processor using openpyxl and xlrd libraries."""

from collections.abc import Iterable

from file_processors.base_processor import BaseFileProcessor


def _format_sheet_rows(rows: Iterable[Iterable[object]]) -> list[str]:
    """Turn raw sheet rows into context-preserving text lines.

    The first row is treated as a header when it has at least two non-empty cells
    and the sheet has more than one row.  Each subsequent data cell is then emitted
    as ``"<header>: <value>"`` so that downstream NER/LLM detection sees the column
    semantics — e.g. a value under an "IBAN" column keeps that association instead of
    being flattened into an undifferentiated bag of words.  One line per row preserves
    record boundaries so entities from different records do not fuse.

    Args:
        rows: Iterable of rows; each row is an iterable of cell values (may be None).

    Returns:
        List of text lines (one per non-empty row), header first when detected.
    """
    norm: list[list[str]] = []
    for row in rows:
        cells = ["" if v is None else str(v).strip() for v in row]
        if any(cells):
            norm.append(cells)

    if not norm:
        return []

    header: list[str] | None = None
    first = norm[0]
    if len(norm) > 1 and sum(1 for c in first if c) >= 2:
        header = first
        data_rows = norm[1:]
    else:
        data_rows = norm

    lines: list[str] = []
    if header is not None:
        lines.append(" | ".join(c for c in header if c))

    for cells in data_rows:
        pairs: list[str] = []
        for i, val in enumerate(cells):
            if not val:
                continue
            if header is not None and i < len(header) and header[i]:
                pairs.append(f"{header[i]}: {val}")
            else:
                pairs.append(val)
        if pairs:
            lines.append(" | ".join(pairs))

    return lines


class XlsxProcessor(BaseFileProcessor):
    """Processor for XLSX (Excel 2007+) files.

    Extracts text from XLSX files using openpyxl library.
    Extracts all cell values from all sheets for PII detection, preserving the
    column-header context of each value (see :func:`_format_sheet_rows`).
    """

    def extract_text(self, file_path: str) -> str:
        """Extract text from an XLSX file.

        Extracts all cell values from all worksheets in the workbook.
        Only extracts actual values, not formulas.

        Args:
            file_path: Path to the XLSX file

        Returns:
            Extracted text content from all cells as a string

        Raises:
            PermissionError: If file cannot be accessed
            FileNotFoundError: If file does not exist
            Exception: For other XLSX processing errors
        """
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise ImportError(
                "openpyxl is required for XLSX processing. "
                "Install it with: pip install openpyxl"
            )

        lines: list[str] = []

        try:
            # Load workbook (read-only mode for better performance)
            workbook = load_workbook(file_path, read_only=True, data_only=True)

            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                lines.extend(_format_sheet_rows(sheet.iter_rows(values_only=True)))

            workbook.close()

        except Exception as e:
            raise Exception(f"Error processing XLSX file: {str(e)}") from e

        return "\n".join(lines)

    @staticmethod
    def can_process(extension: str) -> bool:
        """Check if this processor can handle XLSX files."""
        return extension.lower() == ".xlsx"


class XlsProcessor(BaseFileProcessor):
    """Processor for XLS (Excel 97-2003) files.

    Extracts text from older XLS files using xlrd library.
    Extracts all cell values from all sheets for PII detection, preserving the
    column-header context of each value (see :func:`_format_sheet_rows`).
    """

    def extract_text(self, file_path: str) -> str:
        """Extract text from an XLS file.

        Extracts all cell values from all worksheets in the workbook.

        Args:
            file_path: Path to the XLS file

        Returns:
            Extracted text content from all cells as a string

        Raises:
            PermissionError: If file cannot be accessed
            FileNotFoundError: If file does not exist
            ImportError: If xlrd is not installed
            Exception: For other XLS processing errors
        """
        try:
            import xlrd
        except ImportError:
            raise ImportError(
                "xlrd is required for XLS processing. Install it with: pip install xlrd"
            )

        lines: list[str] = []

        try:
            workbook = xlrd.open_workbook(file_path)

            for sheet_index in range(workbook.nsheets):
                sheet = workbook.sheet_by_index(sheet_index)
                rows = (
                    [sheet.cell(r, c).value for c in range(sheet.ncols)]
                    for r in range(sheet.nrows)
                )
                lines.extend(_format_sheet_rows(rows))

        except Exception as e:
            raise Exception(f"Error processing XLS file: {str(e)}") from e

        return "\n".join(lines)

    @staticmethod
    def can_process(extension: str) -> bool:
        """Check if this processor can handle XLS files."""
        return extension.lower() == ".xls"
